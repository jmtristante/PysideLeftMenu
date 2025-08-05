import re
import pandas as pd
import os
import yaml
from openpyxl import load_workbook
import fnmatch

MODULE_BASE = os.path.dirname(os.path.dirname(__file__))

def get_type(format):
    for regla in config['formats']:
        match = re.match(regla["pattern"], format, flags=re.IGNORECASE)
        if match:
            resultado = {"type": regla["type"]}
            for clave, valor in regla.items():
                if clave in ("pattern", "type"):
                    continue
                # Reemplaza \1, \2, etc. por los grupos de la expresión regular
                resultado[clave] = re.sub(regla["pattern"], valor, format, flags=re.IGNORECASE)
            return resultado
    print(f'Alerta: la definicion "{format}" no cumple ningun patron')
    return None  # No coincide con ningún patrón


def get_dim(format, dim):
    if 'DIM' not in dim:
        return ''
    if 'NUM' in format:
        return f'{dim}.value'
    if 'CHAR' in format:
        return f'{dim}.description'
    return ''


def load_extractor(name, sheet, pk_fields):
    extractor_data_yaml = {"name": name, 'data': []}

    for index, row in sheet.iterrows():
        if index < config['first_line']:
            continue
        field = row.iloc[config['field_column']]
        format = row.iloc[config['format_column']]
        calculation = row.iloc[config['calculation_column']]

        if row.size > config['dimension_column']:
            dim = str(row.iloc[config['dimension_column']])
        else:
            dim = 'nan'

        type = get_type(format)

        if type is None:
            print ('Tipo nulo')

        dim = get_dim(format, dim)

        definition = {
            "name": field,
            "type": type.get("type"),
            "size": int(type.get("size"))
        }
        if type.get("precision"):
            definition["precision"] = int(type.get("precision"))
        if field in pk_fields:
            definition["pk"] = "Y"

        extractor_data_yaml['data'].append(definition)
    return extractor_data_yaml


def create_directory(directory_name):
    # Create the directory
    try:
        os.mkdir(directory_name)
        print(f"Directory '{directory_name}' created successfully.")
    except FileExistsError:
        print(f"Directory '{directory_name}' already exists.")
    except PermissionError:
        print(f"Permission denied: Unable to create '{directory_name}'.")
    except Exception as e:
        print(f"An error occurred: {e}")


def load_pk(sheet_name, sheet):
    pk = []
    for row in sheet.iter_rows(min_row=config['first_line'] + 1, max_row=sheet.max_row):
        field = row[config['field_column']]
        if field.value and field.font and field.font.bold:
            pk.append(field.value)

    pk_text = "\n".join(str(x) for x in pk)
    return pk_text


def read_excel(path):
    excel_data = []
    print(f'Cargando fichero')
    excel = pd.ExcelFile(path)
    print(f'Fichero 1 cargado')
    excel2 = load_workbook(path)  # Este es para sacar las negritas solamente
    print(f'Fichero 2 cargado')
    for sheet_name in excel.sheet_names:
        sheet = pd.read_excel(excel, sheet_name, header=None)
        sheet2 = excel2[sheet_name]

        real_name = sheet_name.replace('_AAMMDD','')
        real_name = real_name.replace('_AAAAMMDD', '')

        # if sheet_name.startswith('HUB') or sheet_name.startswith('LNK') or sheet_name.startswith('SAT'):
        if any(fnmatch.fnmatch(real_name, pattern) for pattern in config['sheets']):
            pk_fields = load_pk(real_name, sheet2)
            extractor_data = load_extractor(real_name, sheet, pk_fields)
            extractor_data['pk'] = load_pk(real_name, sheet2)
            excel_data.append(extractor_data)

    return excel_data


def create_scope(scope, version, excel_data):
    # Crear directorio del scope
    directory_name = os.path.join(MODULE_BASE, scope, version)
    create_directory(directory_name)

    for extractor_data in excel_data:
        name = extractor_data['name']
        data = extractor_data['data']
        pk = extractor_data['pk']

        # Crear directorio del extractor
        ex_directory_name = os.path.join(directory_name, name)
        create_directory(ex_directory_name)

        # Creamos el fichero de definicion
        with open(os.path.join(ex_directory_name, "structure.yaml"), "w") as text_file:
            yaml.dump(data, text_file, allow_unicode=True, sort_keys=False, default_flow_style=False)

        # Creamos el fichero de pk
        #with open(os.path.join(ex_directory_name, "pk.prop"), "w") as text_file:
        #    text_file.write(pk)

def load_config(scope):
    global config
    config_path = os.path.join(MODULE_BASE, scope, "ddr.yaml")
    with open(config_path) as config_file:
        config = yaml.safe_load(config_file)


if __name__ == "__main__":
    # Variables
    scope = 'GRT'
    path = 'GRT.xlsm'
    version = '20250618'

    # Cargar configuracion del DDR
    config = {}
    load_config(scope)

    # Cargar el excel
    excel_data = read_excel(path)

    # Crear los archivos
    create_scope(scope, version, excel_data)
